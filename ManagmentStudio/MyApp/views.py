from django.shortcuts import render, redirect
# from django.utils import timezone
from . import models
from .forms import RegistrationForm
from .models import PersonData
import os, sys
# import xlsxwriter
# import openpyxl
import pandas as pd
from openpyxl import load_workbook


# Create your views here.

def register_page(request):
    context = {
        "form": RegistrationForm}
    return render(request, 'dataRegisterPage.html', context)


def add_person_data(request):
    form = RegistrationForm(request.POST)

    if form.is_valid():
        # print(str(form.data['data']))
        mydata = str(request.POST['data']) + " . "
        locationOfData = GetLocation(mydata.strip())
        typeOfData = GetRequestOfCitizen(mydata.strip())
        record = mydata.strip()
        print(record)
        arg = GetArgument(mydata.strip())
        print(arg)
        # print(str(form.name))
        print(str(form.cleaned_data['name']))
        dt = pd.datetime.now()
        data_frame = pd.DataFrame({'اسم المواطن': [str(form.cleaned_data['name']).strip()],
                                   'رقم الهاتف': [str(form.cleaned_data['phone']).strip()],
                                   'العنوان': [str(form.cleaned_data['address']).strip()],
                                   'تاريخ الشكوى': [dt],
                                   'الشكوى كاملة': [mydata],
                                   'الجهة المعنية': [locationOfData],
                                   'الغرض من الشكوى': [typeOfData],
                                   'الباقي': [arg],
                                   })
        save_in_excel(data_frame)
        myUser = PersonData(name=form.cleaned_data['name'],
                            phone=form.cleaned_data['phone'],
                            address=form.cleaned_data['address'],
                            data=form.cleaned_data['data'],
                            date=dt,
                            locationOfData=locationOfData,
                            typeOfData=typeOfData,
                            argument=arg
                            )
        myUser.save()

        return redirect('confirm', person_name=str(request.POST['name']))
    return redirect('Register')


def confirm(request, person_name):
    personData = models.PersonData.objects.filter(name=person_name)
    context = {
        'name': person_name,
        'personData': personData
    }
    return render(request, 'confirmPage.html', context=context)


def GetLocation(msg):
    earth = ["أيجار", "فرز", "اراضي", "قرض", "القرض", "اسكان", "صندوق", "اسكن", "ايجار", "قطعة", "ارض", "عرصة",
             "بلديه", "عرصه", "تجاوز", "املك", "قطعه","طابو","للقطع","الفرز"
             "مفروزة","الاراضي", "اصحاب", "اراضينا", "قطع", "أرض", "البلدية", "بلدية",
             "سكنية", "مساحة", "قاصرين", "القاصرين"]
    education = ["مدرس", "مدرسة", "مدرسين", "معلمين", "معلمات", "تربية", "مدرسة", "مديرية", "التربية", "معدل", "ناجح",
                 "راسب", "مكمل", "العلمي", "الفرع", "الاحيائي", "معلم", "معلمة", "متوسطة", "اعدادية", "اعداديه",
                 "الابتدائية", "الابتدائيه", "روضة", "رياض", "الاطفال", "مشرف", "تربوي", "تعليم", "مهني", "التعليم",
                "مدارس", "للتربية", "خدمة", "المهني","لمدارس"]
    work = [" بتعييني", "السياسين", "بشمولي", "براتب", "عاجزة", "رعاية", "معيل", "شبكة", "الحماية", "الاجتماعية",
            "هيئة", "ذوي", "الاعاقة", "رعاية", "المراة", "الرعاية", "اجتماعية", "الاحتياج", "الخاص", "شبكة", "الحماية",
            "راتب", "المعيل", "المفصولين", "مفصول", "للتعيين", "وعاطل", "العمل", "عاطل", "عمل", "شركة", "الإعاقة",
            "الراتب", "راتب"]
    health = ["مصابة", "المصاب", "علاج", "مصاب", "مصابين", "اللجنة", "اللجان", "الطبية", "قرار", "طبي", "معاق", "كرسي",
              "متحرك", "مريض", "مريضة", "راقد", "راقده", "المستشفى", "صيدلية", "اعفاء", "دفع", "اجور", "اعفائنا", "دفع",
              "اجور", "المستشفى", "الصيدليات", "بامراض", "عملية", "المصابين", "الصحة", "بمرض", "المريض", "مستشفى",
              "مركز", "صحي", "رعاية", "صحية", "لقاح", "دواء", "ادوية", "طبيب", "طبيبة", "ممرض", "ممرضة", "اسعاف",
              "اسعاف", "سيستر", "دكتور"]
    municipalities = ["الخدمات", "خدمات", "بلديه", "بلدية", "العزيزية", "الصويرة", "الحي", "بدر", "النعمانية", "شراء",
                      "قطعة", "ارض", "تعويض",
                      "متجاوزين", "تجاوز", "تخصيص", "استلام", "عدم", "ارض", "عرصه", "بلدية", "بتخصيص", "تجاوز", "مشاع",
                      "مصفره", "اخلاء", "التجاوز", "ازالة", "منحني", "منحنا", "حصولي", "قطعه", "كشك", "فتح", "تبليط",
                      "تعديل", "الشوارع", "نفايات", "حاوية", "حاويات", "الازبال", "النفايات", "الطمر", "الصحي", "موقع",
                      "مساطحة", "عدم", "استفادة", "محطة", "وسطية", "عامل", "النظافة", "القسم", "البلدي", "انقاض",
                      "مقاطعة", "استثمارها", "استثمار", "الشروط", "الهندسية"]
    oil = ["شركة", "النفط", "شركه", "نفط", "حقل", "الاحدب", "الشركة", "الروسية", "الصينية", "الواحة", "كاز", "بروم",
           "هيئة", "نفط", "واسط", "الاحدب"]
    water = ["بالمياه", "مد", "انقطاع", "المياه", "انقاء", "كسر", "انبوب", "الماء", "مجمع", "ماء", "اسالة", "منظومة",
            "للماء","المي", "صالح", "للشرب", "شرب", "الصالح"]
    sewers = ["مجاري", "طفح", "تصريف", "مياه", "محطة", "اسنه", "مركز"]
    elctric = ["الكهرباء", "انقطاع", "صيانة", "كهرباء", "عدم", "توفر", "التيار", "ضعف", "الخصخصة", "تجاوز", "الشبكة",
               "اعمدة", "أعمدة", "دائرة", "الامبير", "", "المولدات"]

    # myList = msg.split()
    countEarth = 0
    countEdu = 0
    countWork = 0
    countHealth = 0
    countMunicipalities = 0
    countOil = 0
    countWater = 0
    countSewers = 0
    countElect = 0

    listEarth = []
    listEdu = []
    listWork = []
    listHealth = []
    listMunicipalite = []
    listOil = []
    listWater = []
    listSewers = []
    listElect = []
    msg = msg.strip()
    myList = msg.split()

    for i in myList:
        if i in earth:
            countEarth = countEarth + 1
            listEarth.append(i)
        if i in education:
            countEdu = countEdu + 1
            if i.strip() == 'تربية' or i.strip() == 'تربيه':
                countEdu += 5
            listEdu.append(i)
        if i in work:
            countWork = countWork + 1
            if i == "الراتب" or i == "راتب":
                countWork += 5
            listWork.append(i)
        if i in health:
            countHealth = countHealth + 1
            listHealth.append(i)
        if i in municipalities:
            countMunicipalities = countMunicipalities + 1
            if i == "كشك":
                countMunicipalities += 5
            listMunicipalite.append(i)
        if i in oil:
            countOil = countOil + 1
            listOil.append(i)
        if i in water:
            countWater = countWater + 1
            if i == "ماء" or i == "الماء" or i == "المي":
                countWater += 5
            listWater.append(i)
        if i in sewers:
            countSewers = countSewers + 1
            if i == "مجاري":
                countSewers += 5
            listSewers.append(i)
        if i in elctric:
            countElect = countElect + 1
            listElect.append(i)

    print("earth:" + str(countEarth) + str(listEarth))
    print("education:" + str(countEdu) + str(listEdu))
    print("work:" + str(countWork) + str(listWork))
    print("health:" + str(countHealth) + str(listHealth))
    print("municipalities:" + str(countMunicipalities) + str(listMunicipalite))
    print("oil:" + str(countOil) + str(listOil))
    print("water:" + str(countWater) + str(listWater))
    print("sewers:" + str(countSewers) + str(listSewers))
    print("elctric:" + str(countElect) + str(listElect))

    print("#" * 80)
    if countEarth > max(countEdu, countWork, countHealth, countMunicipalities, countOil, countWater, countSewers,
                        countElect):
        print("  countEarth  ".center(80, '#'))
        return "قطع الاراضي"
    elif countEdu > max(countEarth, countWork, countHealth, countMunicipalities, countOil, countWater, countSewers,
                        countElect):
        print("  countEdu   ".center(80, '#'))
        return "التربية"
    elif countWork > max(countEarth, countEdu, countHealth, countMunicipalities, countOil, countWater, countSewers,
                         countElect):
        print("  countWork  ".center(80, '#'))
        return "العمل والشؤون الاجتماعية"
    elif countHealth > max(countEarth, countEdu, countWork, countMunicipalities, countOil, countWater, countSewers,
                           countElect):
        print("  countHealth  ".center(80, '#'))
        return "الصحة"
    elif countMunicipalities > max(countEarth, countEdu, countWork, countHealth, countOil, countWater, countSewers,
                                   countElect):
        print("  countMunicipalities  ".center(80, '#'))
        return "بلديات"
    elif countOil > max(countEarth, countEdu, countWork, countHealth, countMunicipalities, countWater, countSewers,
                        countElect):
        print("  countOil  ".center(80, '#'))
        return "النفط"
    elif countWater > max(countEarth, countEdu, countWork, countHealth, countMunicipalities, countOil, countSewers,
                          countElect):
        print("  countWater  ".center(80, '#'))
        return "الماء"
    elif countSewers > max(countEarth, countEdu, countWork, countHealth, countMunicipalities, countOil, countWater,
                           countElect):
        print("  countSewers  ".center(80, '#'))
        return "المجاري"
    elif countElect > max(countEarth, countEdu, countWork, countHealth, countMunicipalities, countOil, countWater,
                          countSewers):
        print("  countElect  ".center(80, '#'))
        return "الكهرباء"
    else:
        print("cant know the location")
        return "الطلب مبهم غير قادر على تحديد الوجهة"
    print("#" * 80)
    return ()


def GetRequestOfCitizen(msg):
    start_word_type = ["نناشد", "المناشده", "مناشده", "مناشدة", "جئت", "يرجى", "راجين", "الموافقة", "الموافقه", "ارجو",
                       "اروم"]
    end_word_type = ["علمن", "علما", "تضمن", ]

    save_start = None
    save_end = None

    myList = msg.split()

    while save_start is None:
        for i in myList:
            if i in start_word_type:
                save_start = i
                if save_start is not None:
                    print(save_start)
                    break
        if save_start is None:
            return msg

    if save_start == "مناشده" or save_start == "مناشدة" or save_start == "المناشده" or save_start == "نناشد":
        return msg

    if save_start == "راجين":
        msgType = myList[myList.index(save_start):]
        return listToString(msgType)

    while save_end is None:
        for i in myList:
            if i in end_word_type:
                save_end = i
                if save_end is not None:
                    break
        if save_end is None:
            save_end = "."
    print(save_end)

    if save_start == "راجين":
        msgType = myList[myList.index(save_start):]
    elif save_start is not None and save_end is not None:
        msgType = myList[myList.index(save_start):myList.index(save_end)]
    else:
        msgType = "خطا"

    return listToString(msgType)


def GetArgument(msg):
    start_word_arg = ["لاننا", "بسبب", "تعتبر", "لكون", "لكوننا", "حسب", "ما", "يقال", "حينها", "لكونها", "لانه",
                      "لانها", "بدليل", "حيث", "لان", "وذلك", "ذلك"]

    myList = msg.split()
    save_start = None
    while save_start is None:
        for i in myList:
            if i in start_word_arg:
                save_start = i
                if save_start is not None:
                    print(save_start)
                    break
        if save_start is None:
            return "لا يوجد"
        if save_start is not None:
            msgArg = myList[myList.index(save_start):]
            return listToString(msgArg)
    return "لا يوجد"


def listToString(s):
    # initialize an empty string
    str1 = " "

    # return string
    return str1.join(s)


def save_in_excel(data_frame):
    folder_path = os.path.dirname(__file__)
    # read existing file
    reader = pd.read_excel(rf'{folder_path}/data.xlsx')
    if reader.empty:
        writer = pd.ExcelWriter(f'{folder_path}/data.xlsx',
                                engine='openpyxl')  # pylint: disable=abstract-class-instantiated
        # try to open an existing workbook
        writer.book = load_workbook(f'{folder_path}/data.xlsx')
        # copy existing sheets
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)

        # write out the new sheet
        data_frame.to_excel(writer, index=False, header=True, startrow=len(reader) + 1)

        writer.close()
    else:
        writer = pd.ExcelWriter(f'{folder_path}/data.xlsx',
                                engine='openpyxl')  # pylint: disable=abstract-class-instantiated
        # try to open an existing workbook
        writer.book = load_workbook(f'{folder_path}/data.xlsx')
        # copy existing sheets
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)

        # write out the new sheet
        data_frame.to_excel(writer, index=False, header=False, startrow=len(reader) + 1)

        writer.close()
